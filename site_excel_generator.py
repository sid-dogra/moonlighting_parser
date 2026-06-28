import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime
import re
import sys
import calendar

# Consolidated site name mapping (single source of truth)
SITE_MAPPINGS = {
    "ECEP 5PM-11PM (5PM-9PM Weekends/Holidays)": {
        "filename": "ECEP",
        "display": "ECEP"
    },
    "Gramercy Weekday 5-9pm": {
        "filename": "Gramercy Weekday",
        "display": "Gramercy weekday"
    },
    "41st St Weekday 7-9pm": {
        "filename": "41st",
        "display": "41st St"
    },
    "CBI Weekday 5-9pm": {
        "filename": "CBI",
        "display": "CBI"
    },
    "53rd St Weekend 8am-4pm": {
        "filename": "53rd St weekend",
        "display": "53rd St weekend"
    },
    "Gramercy Weekend 9am-5pm": {
        "filename": "Gramercy Weekend",
        "display": "Gramercy weekend"
    },
    "Forest Hills Saturday 8am-4pm": {
        "filename": "Forest Hills",
        "display": "Forest Hills"
    },
    "Garden City Sunday 8am-4pm": {
        "filename": "Garden City",
        "display": "Garden City"
    },
    "Lake Success Saturday 8am-4pm": {
        "filename": "Lake Success",
        "display": "Lake Success"
    },
    "HCC Weekday 5-7 PM": {
        "filename": "HCC Weekday",
        "display": "HCC weekday"
    },
    "Cancer Center Weekend 8:30AM-5PM": {
        "filename": "Cancer Center Weekend",
        "display": "Cancer Center weekend"
    },
    "Bay Ridge 6-9P+1hr trvl (8A-4P Wknd)": {
        "filename": "Bay Ridge",
        "display": "Bay Ridge"
    },
    "Cancer Center weekday 5-8pm": {
        "filename": "Cancer Center Weekday",
        "display": "Cancer Center weekday"
    },
    "FPO Weekday 5-7 PM": {
    "filename": "FPO Weekday",
    "display": "FPO weekday"
    },
    "32nd St Sunday 8AM-4PM": {
        "filename": "32nd St Weekend",
        "display": "32nd St weekend"
    },
    "32nd St Weekend 8AM-4PM": {
        "filename": "32nd St Weekend",
        "display": "32nd St weekend"
    },
    "41st St Sunday 8AM-4PM": {
        "filename": "41st St Weekend",
        "display": "41st St weekend"
    },
    "41st St Weekend 8AM-4PM": {
        "filename": "41st St Weekend",
        "display": "41st St weekend"
    },
    "53rd St Weekday 5-9pm": {
        "filename": "53rd St weekday",
        "display": "53rd St weekday"
    },
    "Bellevue Supershift 5PM-9PM": {
        "filename": "Bellevue Supershift",
        "display": "Bellevue Supershift"
    },
    "CBI Weekend 8am-4pm": {
        "filename": "CBI weekend",
        "display": "CBI weekend"
    },
    "41st St Weekend 4PM-8PM": {
        "filename": "41st St Weekend PM",
        "display": "41st St weekend PM"
    },
    "Greenpoint 8AM-4PM": {
        "filename": "Greenpoint",
        "display": "Greenpoint"
    },
    "Cancer Center Weekend 8A-5:30PM": {
        "filename": "Cancer Center Weekend 8A-5:30PM",
        "display": "Cancer Center weekend 8A-5:30PM"
    },
    "Cancer Center Weekend 8A-4PM": {
        "filename": "Cancer Center Weekend 8A-4PM",
        "display": "Cancer Center weekend 8A-4PM"
    }
}

# Fellow shift site definitions
FELLOW_SITES = {
    "7AM_CBI": {"display": "7AM CBI", "hours": 1},
    "7AM_32nd": {"display": "7AM 32nd", "hours": 1},
    "32nd_Late": {"display": "32nd Late", "hours": 4},
    "FPO_Weekend": {"display": "FPO Weekend", "hours": 8}
}


def is_holiday_name(value):
    """Check if value is a holiday name (all caps text like 'NEW YEAR'S DAY')"""
    if not value or not isinstance(value, str):
        return False
    # Holiday names are typically all caps and longer than 3 chars
    return value.isupper() and len(value.strip()) > 3


def extract_month_from_filename(filename):
    """Extract month name from filename like 'July_moonlighting.xlsx'"""
    basename = os.path.basename(filename)
    month_match = re.match(r'([A-Za-z]+)_moonlighting\.xlsx', basename)
    return month_match.group(1) if month_match else None


def get_month_number(month_name):
    """Convert month name to number (e.g., 'July' -> 7)"""
    try:
        return list(calendar.month_name).index(month_name.capitalize())
    except ValueError:
        return None


def is_date_in_target_month(date_value, target_month_num):
    """Check if a date belongs to the target month"""
    if not date_value:
        return False
    
    try:
        if isinstance(date_value, datetime):
            return date_value.month == target_month_num
        elif isinstance(date_value, str):
            date_obj = datetime.strptime(date_value, "%m/%d/%Y")
            return date_obj.month == target_month_num
    except (ValueError, TypeError):
        pass
    return False


def format_date_for_output(date_value):
    """Format date as 'Saturday, July 5, 2025'"""
    if not date_value:
        return ""
    
    try:
        if isinstance(date_value, str):
            date_obj = datetime.strptime(date_value, "%m/%d/%Y")
        elif isinstance(date_value, datetime):
            date_obj = date_value
        else:
            return str(date_value)
        
        return date_obj.strftime("%A, %B %d, %Y")
    except (ValueError, TypeError):
        return str(date_value)


def format_date_as_sortable(date_value):
    """Format date as MM-DD-YYYY for sorting"""
    if not date_value:
        return ""
    
    try:
        if isinstance(date_value, datetime):
            return date_value.strftime("%m-%d-%Y")
        elif isinstance(date_value, str):
            date_obj = datetime.strptime(date_value, "%m/%d/%Y")
            return date_obj.strftime("%m-%d-%Y")
    except (ValueError, TypeError):
        pass
    return str(date_value)


def load_workbook_with_data(input_file):
    """Load workbook and return worksheet"""
    try:
        wb = load_workbook(input_file, data_only=False)
        return wb.active
    except Exception as e:
        print(f"Error loading workbook: {e}")
        raise


def identify_black_cells(ws):
    """Identify all black cells in the worksheet"""
    black_cells = set()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if (cell.fill and hasattr(cell.fill, 'start_color') and 
                cell.fill.start_color.rgb == 'FF000000'):
                black_cells.add((row, col))
    return black_cells


def identify_sites(ws):
    """Identify site headers and their column ranges"""
    site_info = []
    unmapped_sites = []
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        
        # Check if this is a valid site header
        if (cell.value and isinstance(cell.value, str) and 
            len(cell.value.strip()) > 4 and 
            cell.value.strip().lower() not in ['hr', 'hrs', 'hours']):
            
            site_name = str(cell.value).strip()
            
            # Get mapping or create safe filename
            if site_name in SITE_MAPPINGS:
                mapping = SITE_MAPPINGS[site_name]
                filename = mapping["filename"]
                display_name = mapping["display"]
            else:
                # Create safe filename for unmapped sites
                filename = re.sub(r'[^\w\s-]', '', site_name).strip()
                filename = re.sub(r'[-\s]+', '_', filename)
                display_name = site_name
                unmapped_sites.append(site_name)
            
            site_info.append({
                'name': site_name,
                'filename': filename,
                'display_name': display_name,
                'start_col': col,
                'end_col': col + 2  # 3 columns: date, name, hours
            })
    
    return site_info, unmapped_sites


def extract_site_data(ws, site, target_month_num, black_cells):
    """Extract data for a specific site, filtering by month and excluding black cells"""
    site_data = []
    
    for row in range(2, ws.max_row + 1):
        # Skip rows with black cells
        if any((row, col) in black_cells 
               for col in range(site['start_col'], site['end_col'] + 1)):
            continue
        
        date_cell = ws.cell(row, site['start_col'])
        name_cell = ws.cell(row, site['start_col'] + 1)
        hours_cell = ws.cell(row, site['start_col'] + 2)
        
        # Skip empty rows
        if not (date_cell.value or name_cell.value or hours_cell.value):
            continue
        
        # Filter by target month
        if not is_date_in_target_month(date_cell.value, target_month_num):
            continue
        
        site_data.append({
            'Date': format_date_for_output(date_cell.value),
            'Name': name_cell.value if name_cell.value else "",
            'Hours': hours_cell.value if hours_cell.value else ""
        })
    
    return site_data


def extract_fellow_data(ws, target_month_num):
    """Extract data from Fellow sheet for all shift types"""
    fellow_data = {
        "7AM_CBI": [],
        "7AM_32nd": [],
        "32nd_Late": [],
        "FPO_Weekend": []
    }

    for row in range(2, ws.max_row + 1):
        # 7 AM Shifts: Col 1=Date, Col 3=Location (CBI/32nd), Col 5=Name
        date_7am = ws.cell(row, 1).value
        location_7am = ws.cell(row, 3).value
        name_7am = ws.cell(row, 5).value

        if (date_7am and name_7am and
            is_date_in_target_month(date_7am, target_month_num) and
            not is_holiday_name(name_7am)):

            site_key = "7AM_CBI" if location_7am and "CBI" in str(location_7am) else "7AM_32nd"
            fellow_data[site_key].append({
                'Date': format_date_for_output(date_7am),
                'Name': str(name_7am).strip(),
                'Hours': FELLOW_SITES[site_key]["hours"]
            })

        # 32nd Late Shifts: Col 8=Date, Col 9=Specialty (check for holiday), Col 10=Name
        date_late = ws.cell(row, 8).value
        specialty_late = ws.cell(row, 9).value
        name_late = ws.cell(row, 10).value

        if (date_late and name_late and
            is_date_in_target_month(date_late, target_month_num) and
            not is_holiday_name(str(specialty_late) if specialty_late else "")):

            fellow_data["32nd_Late"].append({
                'Date': format_date_for_output(date_late),
                'Name': str(name_late).strip(),
                'Hours': FELLOW_SITES["32nd_Late"]["hours"]
            })

        # FPO Weekend Shifts: Col 12=Date, Col 14=Name
        date_fpo = ws.cell(row, 12).value
        name_fpo = ws.cell(row, 14).value

        if (date_fpo and name_fpo and
            is_date_in_target_month(date_fpo, target_month_num) and
            not is_holiday_name(name_fpo)):

            fellow_data["FPO_Weekend"].append({
                'Date': format_date_for_output(date_fpo),
                'Name': str(name_fpo).strip(),
                'Hours': FELLOW_SITES["FPO_Weekend"]["hours"]
            })

    return fellow_data


def create_site_excel(site_data, output_file):
    """Create Excel file for a single site"""
    site_wb = Workbook()
    site_ws = site_wb.active
    site_ws.title = "Schedule"
    
    # Add headers with styling
    headers = ['Date', 'Name', 'Hours']
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for col, header in enumerate(headers, start=1):
        cell = site_ws.cell(1, col, header)
        cell.fill = header_fill
    
    # Add data
    for row_idx, row_data in enumerate(site_data, start=2):
        site_ws.cell(row_idx, 1, row_data['Date'])
        site_ws.cell(row_idx, 2, row_data['Name'])
        site_ws.cell(row_idx, 3, row_data['Hours'])
    
    # Auto-adjust column widths
    for column in site_ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        column_letter = column[0].column_letter
        site_ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    site_wb.save(output_file)


def collect_resident_data(ws, site_info, target_month_num, black_cells):
    """Collect all moonlighting data organized by resident"""
    resident_data = {}

    for site in site_info:
        for row in range(2, ws.max_row + 1):
            # Skip rows with black cells
            if any((row, col) in black_cells
                   for col in range(site['start_col'], site['end_col'] + 1)):
                continue

            date_cell = ws.cell(row, site['start_col'])
            name_cell = ws.cell(row, site['start_col'] + 1)
            hours_cell = ws.cell(row, site['start_col'] + 2)

            # Only process complete entries in target month
            if not (date_cell.value and name_cell.value and hours_cell.value and
                    is_date_in_target_month(date_cell.value, target_month_num)):
                continue

            resident_name = str(name_cell.value).strip()
            date_str = format_date_as_sortable(date_cell.value)

            if resident_name not in resident_data:
                resident_data[resident_name] = []

            resident_data[resident_name].append({
                'date': date_str,
                'site': site['display_name'],
                'hours': hours_cell.value
            })

    return resident_data


def collect_fellow_summary_data(fellow_data):
    """Convert fellow site data into trainee summary format"""
    trainee_data = {}

    for site_key, entries in fellow_data.items():
        display_name = FELLOW_SITES[site_key]["display"]

        for entry in entries:
            name = entry['Name']
            if name not in trainee_data:
                trainee_data[name] = []

            # Parse date back to sortable format
            try:
                date_obj = datetime.strptime(entry['Date'], "%A, %B %d, %Y")
                date_str = date_obj.strftime("%m-%d-%Y")
            except (ValueError, TypeError):
                date_str = entry['Date']

            trainee_data[name].append({
                'date': date_str,
                'site': display_name,
                'hours': entry['Hours']
            })

    return trainee_data


def merge_trainee_data(resident_data, fellow_data):
    """Merge resident and fellow data into single trainee dict"""
    merged = dict(resident_data)

    for name, entries in fellow_data.items():
        if name in merged:
            merged[name].extend(entries)
        else:
            merged[name] = entries

    return merged


def generate_trainee_summary(trainee_data, month, output_dir):
    """Generate text summary report of trainee (resident + fellow) moonlighting hours"""
    report_lines = [
        f"{month} Moonlighting Summary",
        "=" * 50,
        ""
    ]

    if not trainee_data:
        report_lines.append("No trainee data found for this month.")
    else:
        # Sort trainees alphabetically
        for trainee_name in sorted(trainee_data.keys()):
            entries = trainee_data[trainee_name]

            # Sort entries by date
            entries.sort(key=lambda x: datetime.strptime(x['date'], "%m-%d-%Y"))

            # Calculate total hours
            total_hours = sum(
                float(entry['hours']) if isinstance(entry['hours'], (int, float)) else 0
                for entry in entries
            )

            # Add trainee section
            report_lines.append(f"{trainee_name}")
            for entry in entries:
                hours_str = f"{entry['hours']:g}" if isinstance(entry['hours'], (int, float)) else str(entry['hours'])
                site_padded = entry['site'].ljust(30)
                report_lines.append(f"  {entry['date']} {site_padded} {hours_str}")

            report_lines.append(f"  {'Total:'.ljust(10)} {total_hours:g}")
            report_lines.append("")

    # Save report
    report_file = os.path.join(output_dir, f"{month}_Trainee_Summary.txt")
    with open(report_file, 'w') as f:
        f.write('\n'.join(report_lines))

    return report_file


def process_moonlighting_schedule(input_file):
    """Main processing function"""
    # Extract and validate month
    month = extract_month_from_filename(input_file)
    if not month:
        print("Error: Could not extract month from filename. Expected format: Month_moonlighting.xlsx")
        return None

    target_month_num = get_month_number(month)
    if not target_month_num:
        print(f"Error: Invalid month name '{month}'. Please use full month names like 'July'.")
        return None

    # Setup output directory
    script_dir = os.path.dirname(os.path.abspath(input_file))
    output_dir = os.path.join(script_dir, month)
    os.makedirs(output_dir, exist_ok=True)

    print(f"Processing file: {input_file}")
    print(f"Detected month: {month} (filtering for month #{target_month_num})")
    print(f"Output directory: {output_dir}")

    # Load workbook and check for multi-sheet format
    wb = load_workbook(input_file, data_only=False)
    sheet_names = wb.sheetnames
    has_fellow_sheet = "Fellow" in sheet_names
    has_resident_sheet = "Resident" in sheet_names

    print(f"Sheets found: {sheet_names}")

    # Determine which sheet to use for resident data
    if has_resident_sheet:
        ws = wb["Resident"]
        print(f"Using 'Resident' sheet for resident data")
    else:
        ws = wb.active
        print(f"Using active sheet (single-sheet format)")

    print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")

    black_cells = identify_black_cells(ws)
    print(f"Found {len(black_cells)} black cells to exclude")

    site_info, unmapped_sites = identify_sites(ws)
    print(f"Found {len(site_info)} resident sites")
    for site in site_info:
        print(f"  - {site['name']} -> {site['filename']}.xlsx")

    # Process each resident site and create Excel files
    sites_with_data = 0
    for site in site_info:
        print(f"\nProcessing site: {site['name']}")
        site_data = extract_site_data(ws, site, target_month_num, black_cells)

        if site_data:
            filename = f"{month}_{site['filename']}.xlsx"
            output_file = os.path.join(output_dir, filename)
            create_site_excel(site_data, output_file)
            print(f"  Created: {output_file} ({len(site_data)} entries)")
            sites_with_data += 1
        else:
            print(f"  No data found for {site['name']}")

    # Collect resident data for summary
    resident_data = collect_resident_data(ws, site_info, target_month_num, black_cells)

    # Process Fellow sheet if it exists
    fellow_summary_data = {}
    fellow_sites_processed = 0
    if has_fellow_sheet:
        print(f"\n--- Processing Fellow Sheet ---")
        fellow_ws = wb["Fellow"]
        print(f"Fellow sheet dimensions: {fellow_ws.max_row} rows x {fellow_ws.max_column} columns")

        fellow_data = extract_fellow_data(fellow_ws, target_month_num)

        # Create Excel files for each fellow site
        for site_key, entries in fellow_data.items():
            if entries:
                filename = f"{month}_{site_key}.xlsx"
                output_file = os.path.join(output_dir, filename)
                create_site_excel(entries, output_file)
                print(f"  Created: {output_file} ({len(entries)} entries)")
                fellow_sites_processed += 1
            else:
                print(f"  No data found for {site_key}")

        # Collect fellow data for summary
        fellow_summary_data = collect_fellow_summary_data(fellow_data)

    # Merge resident and fellow data for combined summary
    trainee_data = merge_trainee_data(resident_data, fellow_summary_data)

    # Generate combined trainee summary
    print(f"\nGenerating trainee summary report...")
    try:
        summary_file = generate_trainee_summary(trainee_data, month, output_dir)
        print(f"Trainee summary created: {summary_file}")
        trainee_count = len(trainee_data)
    except Exception as e:
        print(f"Error generating trainee summary: {e}")
        import traceback
        traceback.print_exc()
        summary_file = "Error generating summary"
        trainee_count = 0

    print(f"\nProcessing complete! Files saved in '{output_dir}' directory")

    return {
        'month': month,
        'sites_processed': sites_with_data + fellow_sites_processed,
        'resident_sites': sites_with_data,
        'fellow_sites': fellow_sites_processed,
        'black_cells_removed': len(black_cells),
        'output_directory': output_dir,
        'unmapped_sites': unmapped_sites,
        'summary_file': summary_file,
        'trainee_count': trainee_count
    }


if __name__ == "__main__":
    print("Moonlighting Schedule Processor")
    print("=" * 50)
    
    # Get input file
    if len(sys.argv) > 1:
        filename = sys.argv[1]
        script_dir = "/Users/siddhantdogra/Documents/Moonlighting 2025-2026"
        input_file = os.path.join(script_dir, filename)
        
        if not os.path.exists(input_file):
            print(f"Error: File not found: {input_file}")
            sys.exit(1)
    else:
        script_dir = "/Users/siddhantdogra/Documents/Moonlighting 2025-2026"
        input_file = os.path.join(script_dir, "July_moonlighting.xlsx")
    
    # Process the file
    try:
        summary = process_moonlighting_schedule(input_file)
        
        if summary is None:
            print("Error: Processing returned no results.")
            sys.exit(1)
        
        print("\n" + "=" * 50)
        print("SUMMARY:")
        print(f"Month: {summary['month']}")
        print(f"Total sites processed: {summary['sites_processed']}")
        print(f"  Resident sites: {summary['resident_sites']}")
        print(f"  Fellow sites: {summary['fellow_sites']}")
        print(f"Black cells removed: {summary['black_cells_removed']}")
        print(f"Trainees found: {summary['trainee_count']}")
        print(f"Output directory: {summary['output_directory']}")
        print(f"Trainee summary: {summary['summary_file']}")

        if summary['unmapped_sites']:
            print(f"\nUnmapped sites found: {len(summary['unmapped_sites'])}")
            for site in summary['unmapped_sites']:
                print(f"  - '{site}'")
        
    except Exception as e:
        print(f"Error processing file: {e}")
        import traceback
        traceback.print_exc()